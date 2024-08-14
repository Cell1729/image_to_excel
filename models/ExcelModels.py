import openpyxl
from openpyxl.styles import PatternFill
from datetime import datetime
import os

from models.ImageModels import ImportImage


class ImportExcel():
    """
    prepare the excel file for further processing
    """
    def __init__(self):
        self.Settingtemplates()
        self.CopyFromtemplates()

    def Settingtemplates(self):
        templates_path = os.path.join(os.path.dirname(__file__), '..', 'templates', 'templates.xlsx')
        self.templates_excel = openpyxl.load_workbook(templates_path)
        self.templates_sheet = self.templates_excel['templates']
    
    def CopyFromtemplates(self):
        now = datetime.now()
        self.filename = f'{now:%Y-%m-%d-%H-%M}'
        outPutExcel = self.templates_excel
        
        output_dir = os.path.join(os.path.dirname(__file__), '..', 'output')
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
        self.output_path = os.path.join(output_dir, f'{self.filename}.xlsx')
        
        self.templates_excel.save(self.output_path)
    
    def GetExcel(self):
        return self.filename, self.output_path
        
class EditExcel():
    def __init__(self, image_path):
        self.image_aspect, self.pixel_info = ImportImage(image_path).GetPixelValue()
        self.setting()
        self.fillColor()
    
    def setting(self):
        self.filename, self.Excel_path = ImportExcel().GetExcel()
        self.OutPutExcel = openpyxl.load_workbook(self.Excel_path)
        self.OutPutSheet = self.OutPutExcel['templates']
    
    def fillColor(self):
        try:
            if self.image_aspect[0] < 1048576 and self.image_aspect[1] < 16384:
                for x in range(0, self.image_aspect[0]):
                    for y in range(0, self.image_aspect[1]):
                        color = self.pixel_info[x, y]
                        hex_color = f'{color[0]:02x}{color[1]:02x}{color[2]:02x}'
                        self.OutPutSheet.cell(row=x+1, column=y+1).fill = PatternFill(patternType='solid', fgColor=hex_color, bgColor=hex_color)
                output_path = os.path.join(os.path.dirname(__file__), '..', 'output', f'{self.filename}.xlsx')
                self.OutPutExcel.save(output_path)
                print('success!!')
            else:
                print(f"Image size is too large to fit in Excel ({self.image_aspect[0]}x{self.image_aspect[1]}).")
        except Exception as e:
            print(f"An error occurred: {e}")
